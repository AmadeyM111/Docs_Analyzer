from openpyxl import load_workbook
from pathlib import Path
import zipfile
import shutil
import mimetypes
import argparse
import logging
import json
import sys
import hashlib
import os
import base64

try:
    from PIL import Image, ExifTags  # type: ignore
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False


logger = logging.getLogger(__name__)


def _estimate_tokens_heuristic(text: str, chars_per_token: float = 3.0) -> int:
    # Rough heuristic: configurable characters per token (default: 3 chars = 1 token)
    if not text:
        return 0
    return max(1, int(len(text) / chars_per_token))


def _make_json_serializable(obj):
    """
    Recursively convert non-JSON-serializable objects (bytes, etc.) to JSON-compatible types.
    """
    if isinstance(obj, bytes):
        # Convert bytes to base64 string for readability
        try:
            return base64.b64encode(obj).decode('utf-8')
        except Exception:
            # Fallback to hex if base64 fails
            return obj.hex()
    elif isinstance(obj, dict):
        return {key: _make_json_serializable(value) for key, value in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [_make_json_serializable(item) for item in obj]
    elif isinstance(obj, (int, float, str, bool, type(None))):
        return obj
    else:
        # For other types, try to convert to string
        try:
            return str(obj)
        except Exception:
            return None


def _compute_sha256(file_path: Path) -> str:
    hasher = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _extract_exif_dict(pil_image) -> dict:
    # Returns a flat EXIF dictionary with human-readable keys
    exif: dict = {}
    try:
        raw = pil_image.getexif()
        if raw:
            tag_map = {ExifTags.TAGS.get(k, k): v for k, v in raw.items()}
            # Small subset of the most useful fields
            exif = {
                "exif_make": tag_map.get("Make"),
                "exif_model": tag_map.get("Model"),
                "exif_date_time_original": tag_map.get("DateTimeOriginal"),
                "exif_orientation": tag_map.get("Orientation"),
            }
            # Full EXIF dump stored under a separate key
            exif["exif_json"] = tag_map
    except Exception as e:
        logger.debug("EXIF извлечь не удалось: %s", e)
    return exif


def inspect_image_file(file_path_str: str) -> dict:
    """
    Return a dictionary with image information.
    Safely handles non-image files (returns only basic fields).
    """
    file_path = Path(file_path_str)
    info: dict = {
        "extracted_path": str(file_path),
        "file_size_bytes": None,
        "sha256": None,
        "mime": None,
        "extension": file_path.suffix.lower(),
        "format": None,
        "color_mode": None,
        "width_px": None,
        "height_px": None,
        "dpi_x": None,
        "dpi_y": None,
    }

    try:
        info["file_size_bytes"] = os.path.getsize(file_path)
    except Exception:
        pass
    try:
        info["sha256"] = _compute_sha256(file_path)
    except Exception:
        pass
    try:
        mime_guess, _ = mimetypes.guess_type(str(file_path))
        info["mime"] = mime_guess
    except Exception:
        pass

    if not PIL_AVAILABLE:
        return info

    try:
        with Image.open(file_path) as im:
            info["format"] = im.format
            info["color_mode"] = im.mode
            width, height = im.size
            info["width_px"] = width
            info["height_px"] = height
            dpi = im.info.get("dpi")
            if isinstance(dpi, tuple) and len(dpi) >= 2:
                info["dpi_x"] = dpi[0]
                info["dpi_y"] = dpi[1]
            # EXIF
            info.update(_extract_exif_dict(im))
    except Exception as e:
        logger.debug("Файл не распознан как изображение или ошибка чтения: %s", e)

    return info


def extract_images_xlsx(xlsx_path, out_dir):
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    logger.info("Начало извлечения изображений из файла: %s", xlsx_path)
    logger.debug("Директория для сохранения: %s", out.resolve())

    # 1) Try via openpyxl API
    wb = load_workbook(xlsx_path, data_only=True)
    results = []
    processed_files = set()  # Track files that have been processed via API

    for worksheet in wb.worksheets:
        logger.info("Обработка листа: %s", worksheet.title)
        images = getattr(worksheet, "_images", [])
        logger.debug("Найдено изображений на листе %s: %d", worksheet.title, len(images))

        for index_in_sheet, image in enumerate(images, start=1):
            # File name and extension
            extension = None
            if hasattr(image, "mime"):
                extension = mimetypes.guess_extension(image.mime) or ".bin"
            elif hasattr(image, "format"):
                extension = f".{image.format.lower()}"
            else:
                extension = ".bin"

            # Anchor address (top-left cell)
            cell_coordinate = None
            if hasattr(image, "anchor") and hasattr(image.anchor, "from"):
                from_anchor = image.anchor.from_
                cell_coordinate = f"{worksheet.cell(row=from_anchor.row+1, column=from_anchor.col+1).coordinate}"
            elif hasattr(image, "anchor") and isinstance(image.anchor, str):
                cell_coordinate = image.anchor
            else:
                cell_coordinate = "unknown"

            # Raw bytes retrieval
            raw_bytes = None
            if hasattr(image, "_data"):
                raw_bytes = image._data()  # bytes

            file_name = f"{worksheet.title}_{cell_coordinate}_{index_in_sheet}{extension}"
            out_file = out / file_name

            if raw_bytes:
                out_file.write_bytes(raw_bytes)
                processed_files.add(out_file.name)  # Track this file
                logger.debug("Сохранено изображение: %s", out_file)
            else:
                logger.warning("Не удалось получить байты изображения для %s", file_name)

            results.append({
                "sheet": worksheet.title,
                "cell": cell_coordinate,
                "file": str(out_file),
                "ok": bool(raw_bytes),
            })

    # 2) Fallback: unpack ZIP and extract media/* (in case API missed something)
    with zipfile.ZipFile(xlsx_path) as archive:
        media_files = [name for name in archive.namelist() if name.startswith("xl/media/")]
        logger.info("Файлов в xl/media/: %d", len(media_files))
        for media_name in media_files:
            target_file_name = Path(media_name).name
            target = out / target_file_name
            existed = target.exists()
            
            # Skip if this file was already processed via openpyxl API
            if target_file_name in processed_files:
                logger.debug("Медиа-файл %s уже обработан через API, пропуск", target_file_name)
                continue
                
            if not existed:
                with archive.open(media_name) as source, open(target, "wb") as destination:
                    shutil.copyfileobj(source, destination)
                logger.debug("Добавлен медиа-файл из архива: %s", target)
            else:
                logger.debug("Медиа-файл уже существует, индексируем: %s", target)
            # Always add to results, even if file already existed
            results.append({"sheet": None, "cell": None, "file": str(target), "ok": True, "existed": existed})

    logger.info("Завершено. Всего записей в результатах: %d", len(results))
    return results


def parse_args(argv):
    parser = argparse.ArgumentParser(description="XLSX analyzer CLI (images extraction and text stats).")
    subparsers = parser.add_subparsers(dest="command")

    # Common option
    parser.add_argument(
        "--log-level",
        default="INFO",
        choices=["CRITICAL", "ERROR", "WARNING", "INFO", "DEBUG"],
        help="Logging level (default: INFO)",
    )

    # Images subcommand (backward compatible defaults)
    images = subparsers.add_parser("images", help="Extract images and generate report")
    images.add_argument("--xlsx", required=False, help="Path to XLSX file")
    images.add_argument("--out", required=False, help="Directory to save extracted images")
    images.add_argument(
        "--json",
        action="store_true",
        help="Print results as JSON to stdout",
    )
    images.add_argument(
        "--out-json",
        help="Save results to the specified JSON file",
    )

    # Text stats subcommand
    text_stats = subparsers.add_parser("text-stats", help="Compute text statistics per sheet and totals")
    text_stats.add_argument("--xlsx", required=True, help="Path to XLSX file")
    text_stats.add_argument(
        "--out-json",
        required=True,
        help="Save text statistics to the specified JSON file",
    )
    text_stats.add_argument(
        "--use-tiktoken",
        action="store_true",
        help="Use tiktoken for token counting (fallback to heuristic if not available)",
    )
    text_stats.add_argument(
        "--encoding",
        default="cl100k_base",
        help="tiktoken encoding name (default: cl100k_base)",
    )
    text_stats.add_argument(
        "--chars-per-token",
        type=float,
        default=3.0,
        help="Characters per token for heuristic estimation (default: 3.0, i.e., 3 chars = 1 token)",
    )

    # If no subcommand provided, maintain backward compatibility with images mode
    # by parsing top-level args as images defaults.
    # Detect legacy pattern: user passed --xlsx/--out at top level
    if argv and not any(arg in ("images", "text-stats") for arg in argv):
        # parse a legacy-style set of flags by temporarily creating a minimal parser
        legacy = argparse.ArgumentParser(add_help=False)
        legacy.add_argument("--xlsx")
        legacy.add_argument("--out")
        legacy.add_argument("--json", action="store_true")
        legacy.add_argument("--out-json")
        known, _ = legacy.parse_known_args(argv)
        if known.xlsx or known.out:
            # inject "images" subcommand
            argv = ["images"] + argv

    return parser.parse_args(argv)


def configure_logging(level_name: str) -> None:
    logging.basicConfig(
        level=getattr(logging, level_name),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )
    logger.debug("Логирование настроено. Уровень: %s", level_name)


if __name__ == "__main__":
    args = parse_args(sys.argv[1:])
    configure_logging(args.log_level)

    if getattr(args, "command", None) == "text-stats":
        # Text statistics workflow
        wb = load_workbook(args.xlsx, data_only=True)
        per_sheet = []
        total_chars = 0
        total_words = 0
        total_cells_with_text = 0
        total_tokens = 0
        # Image bytes from xl/media without extracting
        total_image_files = 0
        total_image_bytes = 0
        total_image_tokens = 0

        encoder = None
        if getattr(args, "use_tiktoken", False):
            try:
                import tiktoken  # type: ignore
                encoder = tiktoken.get_encoding(args.encoding)
            except Exception as e:
                logger.warning("tiktoken unavailable or failed to load (%s); using heuristic.", e)
                encoder = None

        chars_per_token = getattr(args, "chars_per_token", 3.0)

        for ws in wb.worksheets:
            sheet_chars = 0
            sheet_words = 0
            sheet_cells_with_text = 0
            sheet_tokens = 0
            for row in ws.iter_rows(values_only=True):
                for value in row:
                    if isinstance(value, str) and value:
                        sheet_cells_with_text += 1
                        sheet_chars += len(value)
                        sheet_words += len(value.split())
                        if encoder:
                            try:
                                sheet_tokens += len(encoder.encode(value))
                            except Exception:
                                sheet_tokens += _estimate_tokens_heuristic(value, chars_per_token)
                        else:
                            sheet_tokens += _estimate_tokens_heuristic(value, chars_per_token)
            per_sheet.append({
                "sheet": ws.title,
                "num_cells_with_text": sheet_cells_with_text,
                "num_chars": sheet_chars,
                "num_words": sheet_words,
                "est_tokens": sheet_tokens,
            })
            total_cells_with_text += sheet_cells_with_text
            total_chars += sheet_chars
            total_words += sheet_words
            total_tokens += sheet_tokens

        # Process images: convert to base64 and count tokens using the same method as text
        try:
            with zipfile.ZipFile(args.xlsx) as zf:
                media_names = [n for n in zf.namelist() if n.startswith("xl/media/")]
                for name in media_names:
                    try:
                        info = zf.getinfo(name)
                        total_image_files += 1
                        image_bytes = zf.read(name)
                        total_image_bytes += len(image_bytes)
                        
                        # Convert image bytes to base64 string
                        base64_str = base64.b64encode(image_bytes).decode('utf-8')
                        base64_chars = len(base64_str)
                        
                        # Count tokens for base64 string using the same method as text
                        if encoder:
                            try:
                                image_tokens = len(encoder.encode(base64_str))
                            except Exception:
                                image_tokens = _estimate_tokens_heuristic(base64_str, chars_per_token)
                        else:
                            image_tokens = _estimate_tokens_heuristic(base64_str, chars_per_token)
                        
                        total_image_tokens += image_tokens
                        logger.debug("Image %s: %d bytes, %d base64 chars, %d tokens", name, len(image_bytes), base64_chars, image_tokens)
                    except KeyError:
                        continue
                    except Exception as e:
                        logger.warning("Failed to process image %s: %s", name, e)
                        continue
        except Exception as e:
            logger.warning("Failed to inspect xl/media for images: %s", e)

        report = {
            "file": str(args.xlsx),
            "sheets": per_sheet,
            "totals_text": {
                "total_cells_with_text": total_cells_with_text,
                "total_chars": total_chars,
                "total_words": total_words,
                "total_est_tokens": total_tokens,
            },
            "images": {
                "total_image_files": total_image_files,
                "total_image_bytes": total_image_bytes,
                "est_image_tokens": total_image_tokens if total_image_files > 0 else None,
                "token_calculation_method": "base64_encoded_then_tokenized",
            },
            "totals_overall": {
                "total_est_tokens_text": total_tokens,
                "total_est_image_tokens": total_image_tokens,
                "total_est_tokens_combined": (total_tokens + total_image_tokens),
            },
            "token_method": "tiktoken" if encoder else f"heuristic_{chars_per_token}chars_per_token",
        }
        out_json_path = Path(args.out_json)
        out_json_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_json_path, "w", encoding="utf-8") as f:
            json.dump(report, f, ensure_ascii=False, indent=2)
        print(f"Text stats JSON saved: {out_json_path}")
        sys.exit(0)

    # Default: images workflow (backward compatible)
    xlsx_path = getattr(args, "xlsx", None)
    output_dir = getattr(args, "out", None)
    if not xlsx_path or not output_dir:
        print("Usage (images): Docs_Analyzer.py images --xlsx <file.xlsx> --out <export_dir> [--out-json <report.json>] [--json]")
        print("Usage (text):   Docs_Analyzer.py text-stats --xlsx <file.xlsx> --out-json <report.json> [--use-tiktoken] [--encoding cl100k_base] [--chars-per-token 3.0]")
        sys.exit(2)

    results = extract_images_xlsx(xlsx_path, output_dir)

    # Дополнительная инспекция изображений
    enriched_results = []
    for item in results:
        file_path = item.get("file")
        if file_path and Path(file_path).exists():
            meta = inspect_image_file(file_path)
            merged = {**item, **meta}
        else:
            merged = dict(item)
        enriched_results.append(merged)

    total = len(results)
    total_ok = sum(1 for item in results if item.get("ok"))
    total_fail = total - total_ok
    print(f"Итог: найдено {total} элементов, сохранено успешно: {total_ok}, ошибок: {total_fail}")

    # Make results JSON-serializable (convert bytes, etc.)
    serializable_results = _make_json_serializable(enriched_results)

    # Запись в JSON-файл, если указан --out-json
    if getattr(args, "out_json", None):
        out_json_path = Path(args.out_json)
        out_json_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_json_path, "w", encoding="utf-8") as f:
            json.dump(serializable_results, f, ensure_ascii=False, indent=2)
        print(f"JSON сохранён: {out_json_path}")

    if args.json:
        print(json.dumps(serializable_results, ensure_ascii=False, indent=2))
    else:
        # Краткий табличный вывод
        print("sheet\tcell\tfile\tok")
        for item in enriched_results:
            print(f"{item.get('sheet')}\t{item.get('cell')}\t{item.get('file')}\t{item.get('ok')}")